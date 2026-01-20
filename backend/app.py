from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, and_
import csv
import io
import os
from datetime import datetime
from dateutil import parser as date_parser
from functools import lru_cache
import json
from flask_compress import Compress

app = Flask(__name__)
CORS(app)
Compress(app)  # Enable gzip compression for all responses

# Configuration for large file uploads
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SQLALCHEMY_ECHO'] = False

# Database Configuration with optimization
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///fuel_station.db?timeout=10'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 3600,
    'pool_pre_ping': True,
}

db = SQLAlchemy(app)

# Cache for dashboard stats
_stats_cache = {}
_cache_timestamp = 0
CACHE_DURATION = 10 * 60  # 10 minutes (increased from 5)

# Database Model with Indexes
class Transaction(db.Model):
    __tablename__ = 'transactions'
    
    id = db.Column(db.Integer, primary_key=True)
    transaction_id = db.Column(db.String(50), unique=True, nullable=False, index=True)
    date = db.Column(db.Date, nullable=False, index=True)
    time = db.Column(db.Time, nullable=False)
    fuel = db.Column(db.String(20), nullable=False, index=True)
    machine_no = db.Column(db.Integer, nullable=False, index=True)
    nozzle_no = db.Column(db.Integer, nullable=False)
    liters = db.Column(db.Float, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_type = db.Column(db.String(50), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'transaction_id': self.transaction_id,
            'date': self.date.isoformat(),
            'time': str(self.time),
            'fuel': self.fuel,
            'machine_no': self.machine_no,
            'nozzle_no': self.nozzle_no,
            'liters': self.liters,
            'unit_price': self.unit_price,
            'amount': self.amount,
            'payment_type': self.payment_type
        }

# Create database tables
with app.app_context():
    db.create_all()

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy'}), 200

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload CSV or Excel file and store data in database"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
            return jsonify({'error': 'Only CSV and Excel files are supported'}), 400
        
        # Read file
        if file.filename.endswith('.csv'):
            stream = io.TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            rows = list(reader)
        else:
            try:
                import openpyxl
                file.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            except Exception as e:
                return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Validate required columns
        required_columns = ['transaction_id', 'date', 'time', 'fuel', 'machine_no', 
                          'nozzle_no', 'liters', 'unit_price', 'amount', 'payment_type']
        
        if not rows:
            return jsonify({'error': 'File is empty'}), 400
        
        missing_columns = [col for col in required_columns if col not in rows[0]]
        if missing_columns:
            return jsonify({'error': f'Missing columns: {missing_columns}'}), 400
        
        # Process data with batch insertion for performance
        added_count = 0
        skipped_count = 0
        errors = []
        batch = []
        batch_size = 1000  # Insert in batches of 1000
        
        for idx, row in enumerate(rows):
            try:
                # Check if transaction already exists
                txn_id = str(row['transaction_id']).strip()
                existing = Transaction.query.filter_by(transaction_id=txn_id).first()
                if existing:
                    skipped_count += 1
                    continue
                
                # Parse date and time
                date_str = str(row['date']).strip()
                time_str = str(row['time']).strip()
                
                date_obj = date_parser.parse(date_str).date()
                time_obj = date_parser.parse(time_str).time()
                
                # Create transaction
                transaction = Transaction(
                    transaction_id=txn_id,
                    date=date_obj,
                    time=time_obj,
                    fuel=str(row['fuel']).strip(),
                    machine_no=int(float(row['machine_no'])),
                    nozzle_no=int(float(row['nozzle_no'])),
                    liters=float(row['liters']),
                    unit_price=float(row['unit_price']),
                    amount=float(row['amount']),
                    payment_type=str(row['payment_type']).strip()
                )
                
                batch.append(transaction)
                added_count += 1
                
                # Insert batch when it reaches batch_size
                if len(batch) >= batch_size:
                    db.session.add_all(batch)
                    db.session.commit()
                    batch = []
                
            except Exception as e:
                skipped_count += 1
                errors.append(f"Row {idx + 1}: {str(e)}")
        
        # Insert remaining transactions
        if batch:
            db.session.add_all(batch)
            db.session.commit()
        
        return jsonify({
            'message': 'File uploaded successfully',
            'added': added_count,
            'skipped': skipped_count,
            'total_processed': added_count + skipped_count,
            'errors': errors[:10] if errors else None  # Return only first 10 errors
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Upload error: {str(e)}'}), 500

@app.route('/api/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics with optional date range filter (optimized)"""
    try:
        # Get optional date range parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build cache key
        cache_key = f"{start_date}_{end_date}"
        
        # Check cache first (10-minute cache)
        import time
        global _stats_cache, _cache_timestamp, CACHE_DURATION
        if cache_key in _stats_cache and time.time() - _cache_timestamp < CACHE_DURATION:
            return _stats_cache[cache_key], 200
        
        # Build query efficiently
        query = Transaction.query
        
        if start_date:
            start_date_obj = date_parser.parse(start_date).date()
            query = query.filter(Transaction.date >= start_date_obj)
        
        if end_date:
            end_date_obj = date_parser.parse(end_date).date()
            query = query.filter(Transaction.date <= end_date_obj)
        
        # Use database aggregation for faster calculation
        results = query.all()
        
        if not results:
            empty_response = {
                'total_sales': 0,
                'total_liters': 0,
                'petrol_sales': 0,
                'diesel_sales': 0,
                'petrol_liters': 0,
                'diesel_liters': 0,
                'total_transactions': 0,
                'average_transaction_value': 0,
                'payment_methods': {},
                'machines_activity': {},
                'daily_trend': {},
                'hourly_trend': {},
                'fuel_distribution': {}
            }
            return jsonify(empty_response), 200
        
        # Single pass calculation
        total_sales = 0
        total_liters = 0
        petrol_sales = 0
        petrol_liters = 0
        diesel_sales = 0
        diesel_liters = 0
        payment_methods = {}
        machines = {}
        daily_trend = {}
        hourly_trend = {}
        
        for t in results:
            # Basic totals
            total_sales += t.amount
            total_liters += t.liters
            
            # Fuel breakdown
            if t.fuel.lower() == 'petrol':
                petrol_sales += t.amount
                petrol_liters += t.liters
            else:
                diesel_sales += t.amount
                diesel_liters += t.liters
            
            # Payment methods
            payment_methods[t.payment_type] = payment_methods.get(t.payment_type, 0) + 1
            
            # Machine activity
            key = f"Machine {t.machine_no}"
            if key not in machines:
                machines[key] = {'sales': 0, 'liters': 0, 'count': 0}
            machines[key]['sales'] += t.amount
            machines[key]['liters'] += t.liters
            machines[key]['count'] += 1
            
            # Daily trend
            date_str = t.date.isoformat()
            if date_str not in daily_trend:
                daily_trend[date_str] = {'sales': 0, 'liters': 0}
            daily_trend[date_str]['sales'] += t.amount
            daily_trend[date_str]['liters'] += t.liters
            
            # Hourly trend
            hour = t.time.hour
            hour_str = f"{hour:02d}:00"
            if hour_str not in hourly_trend:
                hourly_trend[hour_str] = {'sales': 0, 'liters': 0, 'count': 0}
            hourly_trend[hour_str]['sales'] += t.amount
            hourly_trend[hour_str]['liters'] += t.liters
            hourly_trend[hour_str]['count'] += 1
        
        # Fuel distribution
        fuel_distribution = {
            'Petrol': {'sales': round(petrol_sales, 2), 'liters': round(petrol_liters, 2), 'percentage': round((petrol_sales/total_sales*100) if total_sales > 0 else 0, 2)},
            'Diesel': {'sales': round(diesel_sales, 2), 'liters': round(diesel_liters, 2), 'percentage': round((diesel_sales/total_sales*100) if total_sales > 0 else 0, 2)}
        }
        
        # Build response
        response = jsonify({
            'total_sales': round(total_sales, 2),
            'total_liters': round(total_liters, 2),
            'petrol_sales': round(petrol_sales, 2),
            'diesel_sales': round(diesel_sales, 2),
            'petrol_liters': round(petrol_liters, 2),
            'diesel_liters': round(diesel_liters, 2),
            'total_transactions': len(results),
            'average_transaction_value': round(total_sales / len(results), 2) if results else 0,
            'payment_methods': payment_methods,
            'machines_activity': {k: {kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()} for k, v in machines.items()},
            'daily_trend': {k: {kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()} for k, v in sorted(daily_trend.items())},
            'hourly_trend': {k: {kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()} for k, v in sorted(hourly_trend.items())},
            'fuel_distribution': fuel_distribution
        })
        
        # Cache the response
        _stats_cache[cache_key] = response
        _cache_timestamp = time.time()
        
        return response, 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """Get all transactions with optional filtering"""
    try:
        fuel_filter = request.args.get('fuel')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        query = Transaction.query
        
        if fuel_filter:
            query = query.filter_by(fuel=fuel_filter)
        
        if date_from:
            from datetime import datetime
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Transaction.date >= from_date)
        
        if date_to:
            from datetime import datetime
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Transaction.date <= to_date)
        
        transactions = query.order_by(Transaction.date.desc(), Transaction.time.desc()).all()
        
        return jsonify([t.to_dict() for t in transactions]), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-template', methods=['GET'])
def download_template():
    """Download CSV template"""
    try:
        csv_content = """transaction_id,date,time,fuel,machine_no,nozzle_no,liters,unit_price,amount,payment_type
TXN001,2024-01-15,08:30:00,Petrol,1,1,50.5,95.5,4823.75,Cash
TXN002,2024-01-15,08:45:00,Diesel,1,2,40.0,87.3,3492.0,Card"""
        
        return csv_content, 200, {
            'Content-Disposition': 'attachment; filename=fuel_station_template.csv',
            'Content-Type': 'text/csv'
        }
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Quick database check
        Transaction.query.limit(1).all()
        return jsonify({'status': 'healthy', 'message': 'API is running'}), 200
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500

@app.route('/api/stats-by-date', methods=['GET'])
def get_stats_by_date():
    """Get detailed statistics for a specific date range"""
    try:
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        query = Transaction.query
        
        if date_from:
            from_date = date_parser.parse(date_from).date()
            query = query.filter(Transaction.date >= from_date)
        
        if date_to:
            to_date = date_parser.parse(date_to).date()
            query = query.filter(Transaction.date <= to_date)
        
        transactions = query.all()
        
        if not transactions:
            return jsonify({'error': 'No data found'}), 404
        
        # Aggregate data
        daily_summary = {}
        for t in transactions:
            date_str = t.date.isoformat()
            if date_str not in daily_summary:
                daily_summary[date_str] = {'sales': 0, 'liters': 0}
            daily_summary[date_str]['sales'] += t.amount
            daily_summary[date_str]['liters'] += t.liters
        
        return jsonify({
            'daily_summary': daily_summary,
            'total_records': len(transactions)
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clear-database', methods=['DELETE'])
def clear_database():
    """Delete all data from the database"""
    try:
        # Delete all transactions
        Transaction.query.delete()
        db.session.commit()
        
        # Clear cache
        global _stats_cache, _cache_timestamp
        _stats_cache = {}
        _cache_timestamp = 0
        
        return jsonify({
            'message': 'All data deleted successfully',
            'status': 'success'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete data: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
